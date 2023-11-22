from flask import Flask, render_template, request
import random
import re
import os
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import pickle
import openpyxl
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer, WordNetLemmatizer
from nltk.tokenize import word_tokenize
import enchant
import nltk
from spellchecker import SpellChecker
from flask_pymongo import PyMongo
from bson.objectid import ObjectId

app = Flask(__name__)
app.config['MONGO_URI'] = 'mongodb://localhost:27017/mydatabase'  # Replace with your MongoDB connection URI
mongo = PyMongo(app)

# Load the training data from an Excel file C:\\Python\\updated
data = pd.read_excel("C:\\Users\\prana\\OneDrive\\Desktop\\bbb\\Request Response 50.xlsx")
excel_file = pd.read_excel("C:\\Users\\prana\\OneDrive\\Desktop\\bbb\\DATA_FINAL_SHOPS.xlsx")  # Replace with the actual path to your Excel file
sheet_name = 'Sheet1'  # Replace with the desired sheet name

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(data['Keywords'], data['Response'], test_size=0.2 ,random_state=42)

# Convert the request data into numerical features using TF-IDF vectorization
vectorizer = TfidfVectorizer()
X_train = vectorizer.fit_transform(X_train)
X_test = vectorizer.transform(X_test)

# Train a machine learning model on the vectorized data
model = RandomForestClassifier()
model.fit(X_train, y_train)
y_pred = model.predict(X_test)

# Calculate the metrics
accuracy = accuracy_score(y_test, y_pred)
precision = precision_score(y_test, y_pred, average='weighted')
recall = recall_score(y_test, y_pred, average='weighted')
f1 = f1_score(y_test, y_pred, average='weighted')

# Print the metrics
print("Accuracy:", accuracy)
print("Precision:", precision)
print("Recall:", recall)
print("F1-score:", f1)


# Save the vectorizer and model to disk for later use
with open('vectorizer.pkl', 'wb') as file:
    pickle.dump(vectorizer, file)
with open('model.pkl', 'wb') as file:
    pickle.dump(model, file)

# Load the vectorizer and model from disk
with open('vectorizer.pkl', 'rb') as file:
    vectorizer = pickle.load(file)
with open('model.pkl', 'rb') as file:
    model = pickle.load(file)



def link(rese,url):
    text = rese
    hyperlink = '<a href="{}">{}</a>'.format(url, text)
    return hyperlink




# Define a function to generate a response to user input using the trained model and the Excel file
def generate_response(s):
    if s=="":
        return "Please enter a valid input"
    # Convert the input text to a vector using the vectorizer
    vectorized_text = vectorizer.transform([s])
    # Use the trained model to make a prediction
    prediction = model.predict(vectorized_text)
    # Get the corresponding response from the Excel file
    response = data.loc[data['Keywords'] == s, 'Response'].values
    if len(response) == 0:
        global user_input_list
        user_input_list = []
        return prediction[0], "Sorry.... still under construction."
    else:
        return prediction[0], response[0]


def display_matching_rows(excel_file, sheet_name, target_row_name):
    # Load the Excel file
    print("s: ",target_row_name)
    workbook = openpyxl.load_workbook(excel_file)

    # Select the worksheet
    worksheet = workbook[sheet_name]

    # Iterate over the rows and find matching rows
    matching_rows = []
    ans = '\n Please login into your instagram and then access the link:) \n'
    cnt = 1
    row_cnt = 0
    for row in worksheet.iter_rows(values_only=True):
        row_cnt+=1
        if row[0] == target_row_name:
            row_number = row_cnt
            data=worksheet.cell(row = row_number, column=2).hyperlink.target
            ans = ans + str(cnt) +".) " + row[1] + ' - '  + data + ', ' + os.linesep
            cnt+=1
    return ans

def spell_check_list(text, spell):
    words = text.split()
    corrected_words = []
    for word in words:
        if word and word not in spell:
            spell.word_frequency.add(word)
            corrected_word = spell.correction(word)
            corrected_words.append(corrected_word)
        else:
            corrected_words.append(word)
    corrected_text = ' '.join(corrected_words)
    return corrected_text


def spell_check(text, spell):
    words = text.split()
    corrected_words = []
    for word in words:
        if word is not None and word not in spell:
            corrected_word = spell.correction(word)
            if corrected_word is not None:
                corrected_words.append(corrected_word)
        else:
            corrected_words.append(word)
    if not corrected_words:
        return ""
    corrected_text = ' '.join(corrected_words)
    return corrected_text


# Example usage for appending words not in US Dictionary
l1 = ["women", "men", "kid", "clothing", "accessories", "foot",
      "footwear", "traditional", "western", "sleep", "lounge",
      "sleeplounge", "gym", "jewellery", "hair", "hand", "bags",
      "handbag", "handbags", "clutches", "back", "packs", "backpack","backpacks", 
      "phone", "belts", "caps", "hats", "sun", "glasses", "sunglasses", "sunglass", "watches",
      "sandals", "boots", "flip", "flops", "flipflops", "flipflop", "flats",
      "shoes", "sneakers", "sports", "heels", "slides", "wallets",
      "chains", "rings", "kurti", "sets", "kurtisets", "kurtiset", "kurtis", "long",
      "frocks", "longfrocks", "longfrock", "sarees", "half", "halfsarees", "halfsaree", "blouses",
      "fabrics", "dupattas", "lehangas", "tops", "t shirts", "tshirts", "t-shirts", "skirts", "shorts",
      "shirts", "sweat", "sweatshirts", "sweatshirt", "jackets", "coats", "blazers", "co-ords",
      "dresses", "jump", "suits", "jumpsuits", "jumpsuit", "ethnic", "kurta", "kurtapyjama","kurtapyjamas",
      "pyjama", "dhotis","sherwani", "sherwanis", "ethnicblazers", "ethnicblazer", "polos", "formal", "formalshirts","formalshirt"]

# Create SpellChecker object
spell = SpellChecker()

# Perform spell check
for word in l1:
    spell.word_frequency.add(word)

# Text Cleaning
def clean_text(text):
    if isinstance(text, str):  # check if input is a string
        cleaned_text = text.lower()  # convert text to lowercase
        cleaned_text = re.sub(r'\d+', '', cleaned_text)  # remove numbers
        cleaned_text = re.sub(r'[^\w\s]', '', cleaned_text)  # remove punctuation
        cleaned_text = cleaned_text.strip()  # remove leading and trailing whitespaces
        return cleaned_text
    elif isinstance(text, list):  # check if input is a list
        cleaned_text = []
        for sentence in text:
            cleaned_sentence = sentence.lower()  # convert sentence to lowercase
            cleaned_sentence = re.sub(r'\d+', '', cleaned_sentence)  # remove numbers
            cleaned_sentence = re.sub(r'[^\w\s]', '', cleaned_sentence)  # remove punctuation
            cleaned_sentence = cleaned_sentence.strip()  # remove leading and trailing whitespaces
            cleaned_text.append(cleaned_sentence)
        return cleaned_text
    else:
        raise TypeError('Input must be a string or a list of strings')

# Tokenization
def tokenize(text):
    words = word_tokenize(text) # word tokenization
    return words

# Parts of Speech Tagging
def pos_tagging(words):
    pos_tags = nltk.pos_tag(words) # parts of speech tagging
    return pos_tags

def stem_compound_words(text):
    lemmatizer = WordNetLemmatizer()
    tokens = text.lower().split()
    stemmed_tokens = []

    for token in tokens:
        # Check if the token is a compound word
        if re.match(r'\w+\w', token):
            # Split the compound word into individual components
            components = token.split()
            lemmatized_components = []

            for component in components:
                # Lemmatize each component separately
                lemmatized_component = lemmatizer.lemmatize(component)
                lemmatized_components.append(lemmatized_component)

            # Join the lemmatized components back into a compound word
            stemmed_token = ' '.join(lemmatized_components)
            if stemmed_token.endswith("wear"):
                stemmed_token = stemmed_token[:-(len("wear"))]  # Remove the "wear" suffix along with a space
                lemmatized_words = [lemmatizer.lemmatize(word) for word in stemmed_token.split()]
                stemmed_token = ' '.join(lemmatized_words)
                lemmatized_words = [lemmatizer.lemmatize(word) for word in stemmed_token.split()]
                if stemmed_token == "kurtis":
                    stemmed_token = "kurti"# Change "kurtis" to "kurti"
            if stemmed_token == "kurtis":
                    stemmed_token = "kurti"
        else:
            stemmed_token = lemmatizer.lemmatize(token)
        if(stemmed_token=='woman'):
            stemmed_token = "women"
        stemmed_tokens.append(stemmed_token)
        print("stemmed_tokens in func",stemmed_tokens)
    stemmed_text = ' '.join(stemmed_tokens)
    return stemmed_text


# Define function to preprocess user input and generate response
def nlp_preprocessing(user_input):
    y = 1
    l = ["women","men","kid","clothing","accessory","foot","traditional","ethnic","western","sleep","lounge",  
     "sleeploungewear","sleepwear","loungewear","loungesleepwear","gym","gymwear","sportswear","sport","jewellery",
     "hair","hairaccessories","handbag","hand","bag","clutch","backpack","back","pack","phone","belt","cap","hat",
     "sunglass","sun","glass","watch","sandal","boot","flipflop","flip","flop","flipflops","flat","shoe","sneaker",
     'heel',"slide","wallet","chain","ring","kurtisets","kurtiset","kurti","set","long","frock","longfrocks",
     "longfrock","halfsarees","halfsaree","half","saree","blouse","dupatta","dupattas","fabric","lehanga",
     "lehangas","top","t shirt","tshirt","t-shirt","t-shirts","skirt","short","sweatshirt","sweat","shirt",
     "jacket","coat","blazer","co-ords","dress","jumpsuit","jump","suit","ethnicjewellery","kurta","kurtapyjama",
     "kurtapyjamas","pyjama","dhoti","sherwani","sherwanis","ethnicblazer","polo","formal","formalshirts",
     "formalshirt","jean","pant","legging"]
    while y:
        keywords = []
        #correcting spelling errors
        correct_spell_word = spell_check(user_input,spell)
        print("spellcheck",correct_spell_word,sep=' ')
        #cleaning the text
        
        if correct_spell_word == "":
            return ""

        if correct_spell_word !=  "":
            cleaned_input = clean_text(correct_spell_word)
            print('clean',cleaned_input)
        
        #stemming 
        if cleaned_input != "":
            stemmed_text = stem_compound_words(cleaned_input)
            print("stemming:",stemmed_text,sep=' ')
        
        if stemmed_text != "":
            tokens = tokenize(stemmed_text)

        if len(tokens) != 0:
            pos_tags = pos_tagging(tokens)
        #print("tokens: ",tokens)
        #print(pos_tags)
        # iterate over the tokens and identify relevant keywords

        for token, pos in pos_tags:
            if pos.startswith('N') or pos.startswith('V') or pos.startswith('J'):
                keywords.append(token.lower())
        
        c=[]
        for i in l:
            if i in keywords:
                c.append(i)
        y-=1
        return c


@app.route('/')
def index():
    return render_template('home.html')


@app.route('/loginAc', methods=['POST'])
def loginAc():
    if request.method == 'POST':
        try:
            username = request.form['username']
            password = request.form['password']
            collection = mongo.db.User
            cursor = collection.find_one({ "username":username, "password":password})
            for record in cursor:
                if  'username' == record:
                    print(record)
                    print(cursor[record])
                    if  username == cursor[record]:
                        return render_template('newcbotline.html')
                    else:
                        return render_template('home.html')
        except:
            print("An exception occurred")
            return render_template('home.html',err='0')
    else:
        return render_template('home.html')

@app.route('/registerAc', methods=['POST'])
def registerAc():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        cpassword = request.form['cpassword']
        if password == cpassword:
            data={ "username":username, "password":password, "cpassword":cpassword }
            mongo.db.User.insert_one(data)
            return render_template('home.html')
        else:
            return render_template('home.html',polo='1')
        

    return render_template('register.html')

@app.route('/feedback', methods=['POST'])
def feedback():
    if request.method == 'POST':
        rating1 = request.form['rating1']
        rating = request.form['rating']
        commentText = request.form['commentText']
        data={ "rating1":rating1, "rating":rating, "commentText":commentText }
        mongo.db.feedback.insert_one(data)
        return render_template('newcbotline.html')
    else:
        return render_template('register.html')


user_input_list = []
@app.route('/chat', methods=['POST'])
def chat():
    end_greetings = ['bye', 'goodbye', 'see you later', 'see ya', 'adios']
    greeting_keywords = ["hello", "hi", "greetings", "hey", "nice to meet you", 'hii']
    greetings = ["Hello! How can I assist you today?", "Welcome! How may I help you?",
                 "Hi there! What can I do for you?", "Hii! What brings you here today?",
                 "Greetings! How can I be of service to you?", "Hey! How can I assist you today?",
                 "What can I do for you? I'm here to help.", "Nice to meet you! How can I assist you today?",
                 "Hi! What can I do for you today?", "Hello there! How may I be of assistance to you today?",
                 "Hi..How can I assist you?", "Good to see you! How can I help you today?",
                 "Hey..What can I help you with?", "Hello"]
    global user_input_list
    input_text = request.form['user_input']
    if input_text.lower() in greeting_keywords:
        response = random.choice(greetings)
    elif input_text.lower() in end_greetings:
        response = "Thank you for using our chatbot. Please let us know your experience in the feedback form.\n DO VISIT AGAIN :)"
    else:
        user_input_list.append(input_text)
        input_txt = ' '.join(user_input_list)
        input_txtt = nlp_preprocessing(input_txt)
        print(input_txtt)
        s = ""
        # Generate a response using the trained model and the Excel file
        for i in input_txtt:
            s+=i
            s+=','
        n = len(s)
        if(n==0):
            return "Please enter a valid input"
        print("snew:",s)
        label, response = generate_response(s[:n-1])

    
    if response.count('shops'):
        list_for_shops = ["kid","kurtisets","kurti","set","kurtiset","kurtis","longfrock","long","frock","longfrocks",
                          "saree","halfsaree","half","halfsarees","blouse","fabric","dupatta","dupattas",
                          "lehangas","lehanga","top","t shirt","tshirt","t-shirt","t-shirts","skirt","short","shirt",
                          "sweatshirt","sweat","jacket","coat","blazer","co-ords","dress","jumpsuit","jump","suit",
                          "sleep","jean","pant","legging","lounge","gym","jewellery","ethnic","hair","handbag",
                          "hand","bag","clutch","backpack","back","pack","phone","belt","cap","hat","sunglass",
                          "sun","glass","watch","sandal","boot","flipflops","flipflop","flip","flop","flat","shoe",
                          "sneaker","sport","heel","slide","kurta","kurtapyjama","pyjama","kurtapyjamas","dhoti",
                          "sherwanis","sherwani","polo","formalshirts","formal","formalshirt","wallet","chain","ring"]
        target_row_name = ""
        print("tragets: ",s)
        for word in list_for_shops:
            if word in input_txtt:
                target_row_name+=word
                target_row_name+=','
        n = len(target_row_name)
        print("target_row_name: ",target_row_name)
        target_row_name = target_row_name[:n-1] 
        #target_row_name = s[:n-1]  # Replace with the row name you want to match
        user_input_list = []
        # Call the function to display the matching rows
        #response1 = display_matching_rows("C:\\Python\\updated\\data (1).xlsx", sheet_name, target_row_name)
        response1 = display_matching_rows("C:\\Users\\prana\\OneDrive\\Desktop\\bbb\\DATA_FINAL_SHOPS.xlsx", sheet_name, target_row_name)
        response = response + os.linesep + response1
     # Store the chat history in MongoDB
    chat_entry = {
        'user_input': input_text,
        'response': response
    }
    response_lines = response.splitlines()
    mongo.db.chat_history.insert_one(chat_entry)
    print("response: final: ",response)
   # return response
    # Return the response lines individually
    return response

@app.route('/history')
def history():
    # Retrieve the chat history from MongoDB
    chat_history = mongo.db.chat_history.find()
    return render_template('history.html', chat_history=chat_history)

@app.route('/clear', methods=['GET'])
def clear():
    global user_input_list
    user_input_list = []
    return 'Chat history cleared.'

if __name__ == '__main__':
    app.run(debug=False, port=5002)  # Change the port number to an available on